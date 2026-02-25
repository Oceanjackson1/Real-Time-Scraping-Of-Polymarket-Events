import csv
import json
import os
from dataclasses import asdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import ScraperSnapshot
from config import EXPORT_DIR


class Exporter:
    """Exports scraped data to CSV, JSON, and Excel files."""

    def __init__(self):
        os.makedirs(EXPORT_DIR, exist_ok=True)

    def export_csv(self, snapshot: ScraperSnapshot) -> str:
        filepath = self._generate_filename("polymarket_events_{}.csv")
        rows = []
        for event in snapshot.events:
            top_market = max(event.markets, key=lambda m: m.volume,
                             default=None)
            row = {
                "event_id": event.id,
                "title": event.title,
                "category": event.category,
                "tags": "; ".join(t.label for t in event.tags),
                "num_markets": len(event.markets),
                "total_volume": round(event.volume, 2),
                "volume_24hr": round(event.volume_24hr, 2),
                "liquidity": round(event.liquidity, 2),
                "description": event.description,
                "created_at": event.created_at or "",
                "oracle_type": top_market.oracle_type if top_market else "",
                "oracle_link": top_market.oracle_link if top_market else "",
                "resolved_by": top_market.resolved_by or "" if top_market else "",
                "top_market_question": top_market.question if top_market else "",
                "top_outcome": "",
                "top_price": "",
                "polymarket_url": event.polymarket_url,
                "scraped_at": snapshot.timestamp,
            }
            if top_market and top_market.outcome_prices:
                try:
                    prices = [float(p) for p in top_market.outcome_prices]
                    best_idx = prices.index(max(prices))
                    if best_idx < len(top_market.outcomes):
                        row["top_outcome"] = top_market.outcomes[best_idx]
                    row["top_price"] = round(max(prices), 4)
                except (ValueError, TypeError):
                    pass
            rows.append(row)

        if rows:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
        return filepath

    def export_json(self, snapshot: ScraperSnapshot) -> str:
        filepath = self._generate_filename("polymarket_events_{}.json")
        data = {
            "scraped_at": snapshot.timestamp,
            "total_events": snapshot.total_events,
            "total_markets": snapshot.total_markets,
            "total_volume": round(snapshot.total_volume, 2),
            "fetch_duration_seconds": snapshot.fetch_duration_seconds,
            "categories": {
                cat: [asdict(e) for e in events]
                for cat, events in snapshot.categories.items()
            },
            "events": [asdict(e) for e in snapshot.events],
        }
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return filepath

    def export_excel(self, snapshot: ScraperSnapshot) -> str:
        filepath = self._generate_filename("polymarket_events_{}.xlsx")
        wb = Workbook()

        # Sheet 名使用采集时间戳（Excel sheet 名不能含 : 等特殊字符）
        sheet_ts = snapshot.timestamp.replace(":", "").replace(" ", "_")

        # ── Sheet 1: 事件明细 ──
        ws = wb.active
        ws.title = sheet_ts

        headers = [
            "Event Name",
            "Event Link",
            "Volume (USD)",
            "End Date",
            "Created At",
            "Rules",
            "Background",
            "Oracle Link",
            "Oracle Type",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                                  fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
        thin_border = Border(
            bottom=Side(style="thin", color="D9E2F3"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        row_num = 2
        stripe_fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC",
                                  fill_type="solid")
        url_font = Font(color="0563C1", underline="single")
        money_format = '#,##0.00'
        wrap_alignment = Alignment(vertical="top", wrap_text=True)

        events_sorted = sorted(snapshot.events,
                                key=lambda e: e.volume, reverse=True)

        for event in events_sorted:
            top_market = max(event.markets, key=lambda m: m.volume,
                             default=None)

            oracle_link = ""
            oracle_type = ""
            rules = ""
            if top_market:
                oracle_link = top_market.oracle_link
                oracle_type = top_market.oracle_type
                rules = top_market.description

            row_data = [
                event.title,                       # Event Name
                event.polymarket_url,              # Event Link
                event.volume,                      # Volume (USD)
                event.end_date or "",              # End Date
                event.created_at or "",            # Created At
                rules,                             # Rules (子市场的解析规则)
                event.description,                 # Background (事件盘口背景)
                oracle_link,                       # Oracle Link
                oracle_type,                       # Oracle Type
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = stripe_fill

            ws.cell(row=row_num, column=3).number_format = money_format

            for url_col in [2, 8]:
                c = ws.cell(row=row_num, column=url_col)
                if c.value:
                    c.font = url_font

            for text_col in [6, 7]:
                ws.cell(row=row_num, column=text_col).alignment = wrap_alignment

            row_num += 1

        col_widths = {
            1: 45, 2: 55, 3: 20,
            4: 22, 5: 22,
            6: 60, 7: 60,
            8: 55, 9: 12,
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Sheet 2: 分类汇总 ──
        ws2 = wb.create_sheet("Category_Summary")
        summary_headers = ["Category", "Events", "Markets",
                           "Total Volume (USD)", "24h Volume (USD)",
                           "Liquidity (USD)"]

        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws2.freeze_panes = "A2"

        cat_rows = []
        for cat_name, cat_events in snapshot.categories.items():
            cat_markets = sum(len(e.markets) for e in cat_events)
            cat_vol = sum(e.volume for e in cat_events)
            cat_24h = sum(e.volume_24hr for e in cat_events)
            cat_liq = sum(e.liquidity for e in cat_events)
            cat_rows.append((cat_name, len(cat_events), cat_markets,
                             cat_vol, cat_24h, cat_liq))

        cat_rows.sort(key=lambda r: r[3], reverse=True)

        for r_idx, row_data in enumerate(cat_rows, 2):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = stripe_fill
            for money_col in [4, 5, 6]:
                ws2.cell(row=r_idx, column=money_col).number_format = money_format

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 20
        ws2.column_dimensions["E"].width = 20
        ws2.column_dimensions["F"].width = 20

        # ── Sheet 3: 预言机汇总 ──
        ws_oracle = wb.create_sheet("Oracle_Summary")
        oracle_headers = ["Oracle Type", "Events", "Markets",
                          "Total Volume (USD)", "24h Volume (USD)"]

        for col_idx, header in enumerate(oracle_headers, 1):
            cell = ws_oracle.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws_oracle.freeze_panes = "A2"

        from collections import defaultdict
        oracle_stats = defaultdict(lambda: {
            "events": 0, "markets": 0, "volume": 0.0, "volume_24hr": 0.0,
        })
        for event in snapshot.events:
            top_m = max(event.markets, key=lambda m: m.volume, default=None)
            otype = top_m.oracle_type if top_m else "Unknown"
            oracle_stats[otype]["events"] += 1
            oracle_stats[otype]["markets"] += len(event.markets)
            oracle_stats[otype]["volume"] += event.volume
            oracle_stats[otype]["volume_24hr"] += event.volume_24hr

        oracle_rows = sorted(oracle_stats.items(),
                              key=lambda x: x[1]["volume"], reverse=True)

        for r_idx, (otype, stats) in enumerate(oracle_rows, 2):
            row_data = [otype, stats["events"], stats["markets"],
                        stats["volume"], stats["volume_24hr"]]
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_oracle.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = stripe_fill
            for money_col in [4, 5]:
                ws_oracle.cell(row=r_idx, column=money_col).number_format = money_format

        total_row = len(oracle_rows) + 2
        total_font = Font(bold=True, size=11)
        ws_oracle.cell(row=total_row, column=1, value="Total").font = total_font
        ws_oracle.cell(row=total_row, column=2,
                       value=sum(s["events"] for s in oracle_stats.values())).font = total_font
        ws_oracle.cell(row=total_row, column=3,
                       value=sum(s["markets"] for s in oracle_stats.values())).font = total_font
        ws_oracle.cell(row=total_row, column=4,
                       value=sum(s["volume"] for s in oracle_stats.values()))
        ws_oracle.cell(row=total_row, column=4).font = total_font
        ws_oracle.cell(row=total_row, column=4).number_format = money_format
        ws_oracle.cell(row=total_row, column=5,
                       value=sum(s["volume_24hr"] for s in oracle_stats.values()))
        ws_oracle.cell(row=total_row, column=5).font = total_font
        ws_oracle.cell(row=total_row, column=5).number_format = money_format

        ws_oracle.column_dimensions["A"].width = 16
        ws_oracle.column_dimensions["B"].width = 10
        ws_oracle.column_dimensions["C"].width = 10
        ws_oracle.column_dimensions["D"].width = 22
        ws_oracle.column_dimensions["E"].width = 22

        # ── Sheet 4: 抓取信息 ──
        ws3 = wb.create_sheet("Scrape_Info")
        info_data = [
            ("Scraped At", snapshot.timestamp),
            ("Total Events", snapshot.total_events),
            ("Total Markets", snapshot.total_markets),
            ("Total Volume (USD)", round(snapshot.total_volume, 2)),
            ("Total Categories", len(snapshot.categories)),
            ("Oracle Types", len(oracle_stats)),
            ("Fetch Duration (seconds)", round(snapshot.fetch_duration_seconds, 2)),
        ]
        label_font = Font(bold=True, size=11)
        for r_idx, (label, value) in enumerate(info_data, 1):
            cell_l = ws3.cell(row=r_idx, column=1, value=label)
            cell_l.font = label_font
            ws3.cell(row=r_idx, column=2, value=value)

        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 30

        wb.save(filepath)
        return filepath

    def _generate_filename(self, template: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(EXPORT_DIR, template.format(timestamp))
